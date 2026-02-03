from django.http import HttpResponse
from django.core.exceptions import ValidationError
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
from io import BytesIO
from .models import Asset


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_asset_template(request):
    """
    下载资产数据模板
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "资产数据模板"
    
    # 定义表头
    headers = [
        '设备名称', '设备编号', '期别', '工序', '产线', '设备类型', '位置', 
        '购买日期', '保修到期日', '成本', '当前价值', '状态', '状态详情'
    ]
    
    # 设置表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 调整列宽
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # 添加一些示例数据
    example_data = [
        ['示例设备1', 'EQP001', '一期', 'PL', '1#', 'Equipment', '车间A', 
         '2023-01-01', '2025-01-01', '50000.00', '45000.00', 'Active', '正常'],
        ['示例设备2', 'EQP002', '二期', 'N1', '2-1#', 'Equipment', '车间B', 
         '2023-05-15', '2025-05-15', '75000.00', '70000.00', 'Under Maintenance', '保养中']
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # 保存工作簿到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回Excel文件
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=资产数据模板.xlsx'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def download_filtered_assets(request):
    """
    根据筛选条件下载资产数据
    """
    # 获取筛选参数
    phase_filter = request.data.get('phase', '')  # '', '一期', '二期'
    
    # 构建查询条件
    filters = {}
    if phase_filter:
        filters['phase'] = phase_filter
    
    # 查询符合条件的资产数据
    assets = Asset.objects.filter(**filters).order_by('phase', 'production_line', 'process', 'name')
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "资产数据"
    
    # 定义表头（与模板一致）
    headers = [
        '设备名称', '设备编号', '期别', '工序', '产线', '设备类型', '位置', 
        '购买日期', '保修到期日', '成本', '当前价值', '状态', '状态详情'
    ]
    
    # 设置表头
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # 调整列宽
    for col_num, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15
    
    # 填充数据
    for row_num, asset in enumerate(assets, 2):
        row_data = [
            asset.name or '',
            asset.ref or '',
            asset.phase or '',
            asset.process or '',
            asset.production_line or '',
            asset.asset_type or '',
            asset.location or '',
            asset.purchase_date.strftime('%Y-%m-%d') if asset.purchase_date else '',
            asset.warranty_expiration_date.strftime('%Y-%m-%d') if asset.warranty_expiration_date else '',
            float(asset.cost) if asset.cost else 0.0,
            float(asset.current_value) if asset.current_value else 0.0,
            asset.status or '',
            asset.state or ''
        ]
        
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
    
    # 保存工作簿到内存
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # 返回Excel文件
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=资产数据.xlsx'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_asset_records(request):
    """
    上传Excel文件并导入资产记录
    """
    try:
        if 'file' not in request.FILES:
            return Response({'error': '没有上传文件'}, status=400)
        
        uploaded_file = request.FILES['file']
        
        # 检查文件类型
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持Excel文件(.xlsx, .xls)'}, status=400)
        
        # 使用pandas读取Excel文件
        df = pd.read_excel(BytesIO(uploaded_file.read()))
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # 处理期数字段
                phase_text = row.get('期别', '')
                if pd.notna(phase_text):
                    if isinstance(phase_text, str):
                        if '一' in phase_text or '1' in str(phase_text):
                            phase = '一期'
                        elif '二' in phase_text or '2' in str(phase_text):
                            phase = '二期'
                        else:
                            phase = str(phase_text)
                    else:
                        phase = str(phase_text)
                else:
                    phase = None
                
                # 处理工序字段
                process = row.get('工序', '')
                if pd.notna(process):
                    process = str(process)
                else:
                    process = None
                
                # 处理产线字段
                production_line = row.get('产线', '')
                if pd.notna(production_line):
                    production_line = str(production_line)
                else:
                    production_line = None
                
                # 处理购买日期
                purchase_date = row.get('购买日期', '')
                if pd.notna(purchase_date):
                    try:
                        from datetime import datetime
                        if isinstance(purchase_date, str):
                            purchase_date = datetime.strptime(purchase_date, '%Y-%m-%d')
                        else:
                            purchase_date = pd.to_datetime(purchase_date)
                    except ValueError:
                        purchase_date = None
                else:
                    purchase_date = None
                
                # 处理保修到期日
                warranty_expiration_date = row.get('保修到期日', '')
                if pd.notna(warranty_expiration_date):
                    try:
                        from datetime import datetime
                        if isinstance(warranty_expiration_date, str):
                            warranty_expiration_date = datetime.strptime(warranty_expiration_date, '%Y-%m-%d')
                        else:
                            warranty_expiration_date = pd.to_datetime(warranty_expiration_date)
                    except ValueError:
                        warranty_expiration_date = None
                else:
                    warranty_expiration_date = None
                
                # 处理成本
                cost = row.get('成本', 0)
                if pd.isna(cost) or cost == '':
                    cost = 0
                else:
                    try:
                        cost = float(cost)
                    except (ValueError, TypeError):
                        cost = 0
                
                # 处理当前价值
                current_value = row.get('当前价值', 0)
                if pd.isna(current_value) or current_value == '':
                    current_value = 0
                else:
                    try:
                        current_value = float(current_value)
                    except (ValueError, TypeError):
                        current_value = 0
                
                # 处理其他字段
                name = row.get('设备名称', '')
                ref = row.get('设备编号', '')
                asset_type = row.get('设备类型', 'Equipment')
                location = row.get('位置', '')
                status = row.get('状态', 'Active')
                state = row.get('状态详情', '')
                
                # 创建或更新资产记录
                asset, created = Asset.objects.update_or_create(
                    ref=ref,
                    defaults={
                        'name': str(name) if pd.notna(name) else '',
                        'phase': phase,
                        'process': str(process) if pd.notna(process) else None,
                        'production_line': str(production_line) if pd.notna(production_line) else None,
                        'asset_type': str(asset_type) if pd.notna(asset_type) else 'Equipment',
                        'location': str(location) if pd.notna(location) else '',
                        'purchase_date': purchase_date,
                        'warranty_expiration_date': warranty_expiration_date,
                        'cost': cost,
                        'current_value': current_value,
                        'status': str(status) if pd.notna(status) else 'Active',
                        'state': str(state) if pd.notna(state) else ''
                    }
                )
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"第{index + 2}行处理失败: {str(e)}")
        
        return Response({
            'message': f'成功导入 {imported_count} 条记录',
            'errors': errors
        })
    
    except Exception as e:
        return Response({'error': str(e)}, status=500)