from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view
from .models_task_plan import TaskPlan
from .models import PlantPhase, Process, ProductionLine
from django.contrib.auth.models import User
from django.utils import timezone
import json


def download_task_plan_template(request):
    """
    下载任务计划Excel模板
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "任务计划模板"

    # 定义标题
    headers = [
        '日期', '任务计划', '任务实施人(用户名)', '状态', '完成进度(%)', 
        '期别', '工序', '产线'
    ]
    
    # 设置标题行样式
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 设置列宽
    for col_num in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    # 添加示例行
    sample_row = [
        '2024-12-31',  # 日期
        '设备保养任务',  # 任务计划
        'admin,user1',  # 任务实施人(多个用户用逗号分隔)
        'pending',      # 状态(pending待处理, in_progress进行中, completed已完成, cancelled已取消)
        '0',            # 完成进度
        '一期',         # 期别
        'N1',           # 工序
        '1#'            # 产线
    ]

    for col_num, value in enumerate(sample_row, 1):
        cell = ws.cell(row=2, column=col_num, value=value)
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # 创建HTTP响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=任务计划模板.xlsx'
    
    wb.save(response)
    return response


@api_view(['POST'])
def download_filtered_task_plans(request):
    """
    下载过滤后的任务计划数据（支持月份筛选）
    """
    from django.http import JsonResponse
    from django.db.models import Q
    
    # 从前端获取过滤条件
    filters = request.data if hasattr(request, 'data') else {}
    month_filter = filters.get('month', '')  # 月份筛选参数，格式为 'YYYY-MM' 或 'all'

    wb = Workbook()
    ws = wb.active
    ws.title = "任务计划数据"

    # 定义标题
    headers = [
        'ID', '日期', '任务计划', '任务实施人', '计划人数', '状态', '完成进度(%)', 
        '期别', '工序', '产线', '创建时间', '更新时间'
    ]
    
    # 设置标题行样式
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # 设置列宽
    column_widths = [10, 15, 30, 20, 10, 12, 15, 10, 10, 10, 20, 20]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # 获取任务计划数据
    task_plans = TaskPlan.objects.all().select_related('phase', 'process', 'production_line').prefetch_related('assigned_users')

    # 根据月份进行筛选
    if month_filter and month_filter != 'all':
        try:
            year, month = month_filter.split('-')
            year = int(year)
            month = int(month)
            
            # 筛选日期在指定月份的任务计划
            task_plans = task_plans.filter(
                Q(date__isnull=False) & 
                Q(date__year=year, date__month=month)
            )
            
        except ValueError:
            # 如果月份格式不正确，返回错误
            response_content = json.dumps({'error': f'月份格式不正确，请使用 YYYY-MM 格式'}, ensure_ascii=False)
            return HttpResponse(response_content, content_type='application/json', status=400)

    # 填充数据
    for row_num, task_plan in enumerate(task_plans, 2):
        try:
            # 获取任务实施人用户名
            assigned_users_str = ','.join([user.username for user in task_plan.assigned_users.all()])
            
            # 获取关联对象名称，添加更安全的空值检查
            phase_name = task_plan.phase.name if task_plan.phase and hasattr(task_plan.phase, 'name') else ''
            process_name = task_plan.process.name if task_plan.process and hasattr(task_plan.process, 'name') else ''
            production_line_name = task_plan.production_line.name if task_plan.production_line and hasattr(task_plan.production_line, 'name') else ''

            row_data = [
                str(task_plan.uuid) if task_plan.uuid else '',  # ID
                task_plan.date.strftime('%Y-%m-%d') if task_plan.date else '',  # 日期
                task_plan.task_description or '',  # 任务计划
                assigned_users_str,  # 任务实施人
                task_plan.planned_people_count or 0,  # 计划人数
                dict(TaskPlan.STATUS_CHOICES).get(task_plan.status, task_plan.status) if task_plan.status else 'unknown',  # 状态
                task_plan.progress or 0.0,  # 完成进度
                phase_name,  # 期别
                process_name,  # 工序
                production_line_name,  # 产线
                task_plan.created_at.strftime('%Y-%m-%d %H:%M:%S') if task_plan.created_at else '',  # 创建时间
                task_plan.updated_at.strftime('%Y-%m-%d %H:%M:%S') if task_plan.updated_at else '',  # 更新时间
            ]

            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else '')
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        except AttributeError as e:
            # 如果出现属性错误，跳过这条记录并记录错误
            print(f"导出任务计划时出现错误，ID: {getattr(task_plan, 'uuid', 'Unknown')}, 错误: {str(e)}")
            continue
        except Exception as e:
            # 记录其他错误
            print(f"导出任务计划时出现未知错误，ID: {getattr(task_plan, 'uuid', 'Unknown')}, 错误: {str(e)}")
            continue

    # 创建HTTP响应
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=任务计划数据.xlsx'
    
    wb.save(response)
    return response


def upload_task_plan_records(file):
    """
    上传任务计划Excel文件并保存到数据库
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(filename=file)
    ws = wb.active
    
    results = {'success_count': 0, 'error_count': 0, 'errors': []}
    
    # 读取Excel数据
    headers = [cell.value for cell in ws[1]]
    
    # 找到各列索引
    header_indices = {}
    for idx, header in enumerate(headers):
        if header:
            header_indices[header.strip()] = idx
    
    # 必需列
    required_columns = ['日期', '任务计划', '任务实施人(用户名)', '状态', '完成进度(%)', '期别', '工序', '产线']
    
    for col in required_columns:
        if col not in header_indices:
            results['errors'].append(f'缺少必需列: {col}')
            results['error_count'] += 1
            return results

    # 从第2行开始读取数据
    for row_idx in range(2, ws.max_row + 1):
        try:
            # 读取数据
            date_val = ws.cell(row=row_idx, column=header_indices['日期'] + 1).value
            task_description = ws.cell(row=row_idx, column=header_indices['任务计划'] + 1).value
            assigned_users_str = ws.cell(row=row_idx, column=header_indices['任务实施人(用户名)'] + 1).value
            status = ws.cell(row=row_idx, column=header_indices['状态'] + 1).value
            progress = ws.cell(row=row_idx, column=header_indices['完成进度(%)'] + 1).value
            phase_name = ws.cell(row=row_idx, column=header_indices['期别'] + 1).value
            process_name = ws.cell(row=row_idx, column=header_indices['工序'] + 1).value
            production_line_name = ws.cell(row=row_idx, column=header_indices['产线'] + 1).value

            # 数据验证
            if not all([date_val, task_description, assigned_users_str]):
                results['errors'].append(f'第{row_idx}行数据不完整')
                results['error_count'] += 1
                continue

            # 处理日期
            if isinstance(date_val, str):
                from datetime import datetime
                try:
                    date_obj = datetime.strptime(date_val, '%Y-%m-%d').date()
                except ValueError:
                    results['errors'].append(f'第{row_idx}行日期格式错误: {date_val}')
                    results['error_count'] += 1
                    continue
            else:
                date_obj = date_val

            # 处理进度
            try:
                progress_val = float(progress) if progress is not None else 0.0
            except (ValueError, TypeError):
                progress_val = 0.0

            # 处理状态
            if status not in dict(TaskPlan.STATUS_CHOICES):
                status = 'pending'  # 默认为待处理

            # 获取关联对象
            phase = None
            if phase_name:
                phase, created = PlantPhase.objects.get_or_create(
                    name=phase_name,
                    defaults={'code': phase_name[:10]}
                )

            process = None
            if process_name:
                process, created = Process.objects.get_or_create(
                    name=process_name,
                    defaults={'code': process_name[:10]}
                )

            production_line = None
            if production_line_name:
                production_line, created = ProductionLine.objects.get_or_create(
                    name=production_line_name,
                    defaults={'code': production_line_name[:10]}
                )

            # 获取用户
            assigned_users = []
            if assigned_users_str:
                user_names = [name.strip() for name in str(assigned_users_str).split(',')]
                for username in user_names:
                    try:
                        user = User.objects.get(username=username)
                        assigned_users.append(user)
                    except User.DoesNotExist:
                        results['errors'].append(f'第{row_idx}行找不到用户: {username}')
                        results['error_count'] += 1
                        continue

            if not assigned_users:
                results['errors'].append(f'第{row_idx}行没有有效的用户')
                results['error_count'] += 1
                continue

            # 创建或更新任务计划
            task_plan = TaskPlan.objects.create(
                date=date_obj,
                task_description=task_description,
                status=status,
                progress=progress_val,
                phase=phase,
                process=process,
                production_line=production_line
            )

            # 分配用户
            task_plan.assigned_users.set(assigned_users)

            results['success_count'] += 1

        except Exception as e:
            results['errors'].append(f'第{row_idx}行处理错误: {str(e)}')
            results['error_count'] += 1

    return results