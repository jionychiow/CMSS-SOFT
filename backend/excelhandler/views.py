import pandas as pd
import json
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import JSONParser
from .models import ExcelUpload, MaintenanceRecord, Equipment, MaintenanceHistory
from django.core.files.base import ContentFile
import os
from django.contrib.auth.models import User
from datetime import datetime


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def upload_excel(request):
    """
    上传Excel文件并解析内容，自动导入维修记录
    """
    if 'file' not in request.FILES:
        return JsonResponse({'error': '没有找到上传的文件'}, status=400)
    
    uploaded_file = request.FILES['file']
    
    # 验证文件类型
    if not uploaded_file.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({'error': '只支持上传Excel文件(.xlsx, .xls)'}, status=400)
    
    try:
        # 读取Excel文件
        df = pd.read_excel(uploaded_file, sheet_name=0)  # 读取第一个工作表
        
        # 转换为字典列表
        data = df.to_dict(orient='records')
        
        # 获取列名
        columns = df.columns.tolist()
        
        # 创建ExcelUpload记录
        excel_record = ExcelUpload.objects.create(
            file=uploaded_file,
            filename=uploaded_file.name,
            uploaded_by=request.user.username
        )
        
        # 尝试导入维修记录数据
        imported_count = import_maintenance_records_from_dataframe(df, excel_record, request.user)
        
        return JsonResponse({
            'message': f'文件上传成功，已导入 {imported_count} 条维修记录',
            'columns': columns,
            'data': data,
            'record_id': excel_record.id,
            'filename': uploaded_file.name,
            'imported_count': imported_count
        })
        
    except Exception as e:
        return JsonResponse({'error': f'处理Excel文件时出错: {str(e)}'}, status=500)


def import_maintenance_records_from_dataframe(df, excel_record, user):
    """
    从DataFrame导入维修记录
    """
    imported_count = 0
    
    # 确定列映射（根据Excel文件分析结果）
    column_mapping = {
        '序号': 'record_number',
        '月份': 'month',
        '产线': 'production_line',
        '工序': 'process',
        '设备或工装\n名称': 'equipment_name',
        '设备编号': 'equipment_code',
        '设备零\n部件\n/部位': 'equipment_part',
        '变更\n原因': 'change_reason',
        '变更前(现状）': 'status_before',
        '变更后（变了什么）': 'status_after',
        '开始日期\n及时间': 'start_datetime',
        '结束日期\n及时间': 'end_datetime',
        '耗用\n时长': 'duration',
        '零件耗材': 'parts_consumables',
        '实施人': 'implementer',
        '确认人': 'confirm_person',
        '验收人': 'acceptor',
        '效果评价': 'effect_evaluation',
        '评价人': 'evaluator',
        '备注': 'remarks',
        'Unnamed: 1': 'shift_type',  # 根据分析结果，班次信息可能在其他列
    }
    
    for index, row in df.iterrows():
        try:
            # 跳过空行
            if pd.isna(row.iloc[1]) and pd.isna(row.iloc[2]):  # 如果序号和月份都是空的，跳过
                continue
            
            # 检查是否包含有效数据
            if pd.isna(row.iloc[1]) or str(row.iloc[1]) == '生产一分部倒班交接班记录表' or str(row.iloc[1]) == '生产一分部长白班交接班记录表':
                continue
                
            # 检查是否有序号
            record_number = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else f"R{index}"
            
            # 检查是否已存在相同记录号
            if MaintenanceRecord.objects.filter(record_number=record_number).exists():
                continue
            
            # 创建或获取设备
            equipment_name = str(row.iloc[4]) if not pd.isna(row.iloc[4]) else "未知设备"
            equipment_code = str(row.iloc[5]) if not pd.isna(row.iloc[5]) else f"EQ{index}"
            
            equipment, created = Equipment.objects.get_or_create(
                code=equipment_code,
                defaults={
                    'name': equipment_name,
                    'production_line': str(row.iloc[2]) if not pd.isna(row.iloc[2]) else "未知产线",
                    'process': str(row.iloc[3]) if not pd.isna(row.iloc[3]) else "未知工序",
                    'part_details': str(row.iloc[6]) if not pd.isna(row.iloc[6]) else None
                }
            )
            
            # 处理时间字段
            start_time = row.iloc[10] if not pd.isna(row.iloc[10]) else None
            end_time = row.iloc[11] if not pd.isna(row.iloc[11]) else None
            
            # 尝试解析时间
            if start_time:
                if isinstance(start_time, str):
                    try:
                        start_time = datetime.fromisoformat(start_time.replace('T', ' '))
                    except:
                        start_time = timezone.now()
                elif pd.isna(start_time):
                    start_time = timezone.now()
            else:
                start_time = timezone.now()
                
            if end_time:
                if isinstance(end_time, str):
                    try:
                        end_time = datetime.fromisoformat(end_time.replace('T', ' '))
                    except:
                        end_time = start_time
                elif pd.isna(end_time):
                    end_time = start_time
            else:
                end_time = start_time
            
            # 处理变更原因
            change_reason = str(row.iloc[7]) if not pd.isna(row.iloc[7]) else "其他"
            # 简化变更原因以匹配预定义选项
            if "故障" in change_reason:
                change_reason = "故障维修"
            elif "保养" in change_reason:
                change_reason = "定期保养"
            elif "改造" in change_reason:
                change_reason = "设备改造"
            elif "更换" in change_reason:
                change_reason = "零件更换"
            else:
                change_reason = "其他"
            
            # 确定班次类型
            shift_type = "其他"  # 默认值
            # 这里可以根据具体业务逻辑来确定班次
            
            # 创建维修记录
            maintenance_record = MaintenanceRecord.objects.create(
                record_number=record_number,
                month=str(row.iloc[2]) if not pd.isna(row.iloc[2]) else "未知月份",
                production_line=str(row.iloc[2]) if not pd.isna(row.iloc[2]) else "未知产线",
                process=str(row.iloc[3]) if not pd.isna(row.iloc[3]) else "未知工序",
                equipment_name=equipment_name,
                equipment_code=equipment_code,
                equipment_part=str(row.iloc[6]) if not pd.isna(row.iloc[6]) else "未知部件",
                change_reason=change_reason,
                status_before=str(row.iloc[8]) if not pd.isna(row.iloc[8]) else "未知状态",
                status_after=str(row.iloc[9]) if not pd.isna(row.iloc[9]) else "未知状态",
                start_datetime=start_time,
                end_datetime=end_time,
                duration=None,  # 需要根据实际数据计算
                parts_consumables=str(row.iloc[14]) if not pd.isna(row.iloc[14]) else None,
                implementer=str(row.iloc[15]) if not pd.isna(row.iloc[15]) else "未知实施人",
                confirm_person=str(row.iloc[16]) if not pd.isna(row.iloc[16]) else None,
                acceptor=str(row.iloc[17]) if not pd.isna(row.iloc[17]) else None,
                effect_evaluation=str(row.iloc[18]) if not pd.isna(row.iloc[18]) else None,
                evaluator=str(row.iloc[19]) if not pd.isna(row.iloc[19]) else None,
                remarks=str(row.iloc[20]) if not pd.isna(row.iloc[20]) else None,
                shift_type=shift_type,
                created_by=user,
                imported_from_excel=True,
                excel_upload=excel_record
            )
            
            # 创建维修历史记录
            MaintenanceHistory.objects.create(
                equipment=equipment,
                maintenance_record=maintenance_record
            )
            
            imported_count += 1
            
        except Exception as e:
            print(f"导入第 {index} 行时出错: {str(e)}")
            continue
    
    return imported_count


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def download_excel_template(request):
    """
    下载Excel模板
    """
    try:
        # 创建一个简单的Excel模板
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from django.http import HttpResponse
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "维修记录模板"
        
        # 添加表头
        headers = ["序号", "月份", "产线", "工序", "设备或工装\n名称", "设备编号", "设备零\n部件\n/部位", 
                  "变更\n原因", "变更前(现状）", "变更后（变了什么）", "开始日期\n及时间", "结束日期\n及时间", 
                  "耗用\n时长", "列1", "零件耗材", "实施人", "确认人", "验收人", "效果评价", "评价人", "备注"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        
        # 调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 将工作簿保存到内存
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=维修记录模板.xlsx'
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'生成Excel模板时出错: {str(e)}'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def export_data_to_excel(request):
    """
    将数据导出为Excel文件
    """
    try:
        # 从前端接收数据
        data = request.data.get('data', [])
        filename = request.data.get('filename', 'export_data.xlsx')
        sheet_name = request.data.get('sheet_name', 'Sheet1')
        
        if not data:
            return JsonResponse({'error': '没有数据可导出'}, status=400)
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 创建Excel文件
        import openpyxl
        from openpyxl.utils.dataframe import dataframe_to_rows
        from django.http import HttpResponse
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # 添加数据到工作表
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 格式化表头
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal="center")
        
        # 自动调整列宽
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # 将工作簿保存到内存
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        return response
        
    except Exception as e:
        return JsonResponse({'error': f'导出Excel文件时出错: {str(e)}'}, status=500)


def read_excel_file(file_path):
    """
    读取Excel文件的辅助函数
    """
    try:
        df = pd.read_excel(file_path, sheet_name=None)  # 读取所有工作表
        result = {}
        for sheet_name, sheet_df in df.items():
            result[sheet_name] = {
                'columns': sheet_df.columns.tolist(),
                'data': sheet_df.to_dict(orient='records'),
                'shape': sheet_df.shape
            }
        return result
    except Exception as e:
        raise Exception(f"读取Excel文件失败: {str(e)}")


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def read_uploaded_excel(request, record_id):
    """
    读取已上传的Excel文件内容
    """
    try:
        excel_record = ExcelUpload.objects.get(id=record_id)
        file_path = excel_record.file.path
        
        # 使用pandas读取Excel文件
        excel_data = read_excel_file(file_path)
        
        return JsonResponse({
            'filename': excel_record.filename,
            'uploaded_at': excel_record.uploaded_at,
            'excel_data': excel_data
        })
        
    except ExcelUpload.DoesNotExist:
        return JsonResponse({'error': '找不到指定的Excel文件记录'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'读取Excel文件时出错: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_maintenance_records(request):
    """
    获取维修记录列表
    """
    try:
        records = MaintenanceRecord.objects.all().order_by('-created_at')
        
        # 转换为字典列表
        records_data = []
        for record in records:
            records_data.append({
                'id': record.id,
                'record_number': record.record_number,
                'month': record.month,
                'production_line': record.production_line,
                'process': record.process,
                'equipment_name': record.equipment_name,
                'equipment_code': record.equipment_code,
                'equipment_part': record.equipment_part,
                'change_reason': record.change_reason,
                'status_before': record.status_before,
                'status_after': record.status_after,
                'start_datetime': record.start_datetime,
                'end_datetime': record.end_datetime,
                'duration': str(record.duration) if record.duration else None,
                'parts_consumables': record.parts_consumables,
                'implementer': record.implementer,
                'confirm_person': record.confirm_person,
                'acceptor': record.acceptor,
                'effect_evaluation': record.effect_evaluation,
                'evaluator': record.evaluator,
                'remarks': record.remarks,
                'shift_type': record.shift_type,
                'created_at': record.created_at,
                'imported_from_excel': record.imported_from_excel
            })
        
        return JsonResponse({
            'records': records_data,
            'count': len(records_data)
        })
        
    except Exception as e:
        return JsonResponse({'error': f'获取维修记录时出错: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_maintenance_record(request, record_id):
    """
    获取单个维修记录详情
    """
    try:
        record = MaintenanceRecord.objects.get(id=record_id)
        
        record_data = {
            'id': record.id,
            'record_number': record.record_number,
            'month': record.month,
            'production_line': record.production_line,
            'process': record.process,
            'equipment_name': record.equipment_name,
            'equipment_code': record.equipment_code,
            'equipment_part': record.equipment_part,
            'change_reason': record.change_reason,
            'status_before': record.status_before,
            'status_after': record.status_after,
            'start_datetime': record.start_datetime,
            'end_datetime': record.end_datetime,
            'duration': str(record.duration) if record.duration else None,
            'parts_consumables': record.parts_consumables,
            'implementer': record.implementer,
            'confirm_person': record.confirm_person,
            'acceptor': record.acceptor,
            'effect_evaluation': record.effect_evaluation,
            'evaluator': record.evaluator,
            'remarks': record.remarks,
            'shift_type': record.shift_type,
            'created_at': record.created_at,
            'updated_at': record.updated_at,
            'imported_from_excel': record.imported_from_excel
        }
        
        return JsonResponse(record_data)
        
    except MaintenanceRecord.DoesNotExist:
        return JsonResponse({'error': '找不到指定的维修记录'}, status=404)
    except Exception as e:
        return JsonResponse({'error': f'获取维修记录时出错: {str(e)}'}, status=500)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def create_maintenance_record(request):
    """
    创建新的维修记录
    """
    try:
        data = request.data
        
        # 创建维修记录
        record = MaintenanceRecord.objects.create(
            record_number=data.get('record_number'),
            month=data.get('month', ''),
            production_line=data.get('production_line', ''),
            process=data.get('process', ''),
            equipment_name=data.get('equipment_name', ''),
            equipment_code=data.get('equipment_code', ''),
            equipment_part=data.get('equipment_part', ''),
            change_reason=data.get('change_reason', '其他'),
            status_before=data.get('status_before', ''),
            status_after=data.get('status_after', ''),
            start_datetime=data.get('start_datetime'),
            end_datetime=data.get('end_datetime'),
            duration=data.get('duration'),
            parts_consumables=data.get('parts_consumables'),
            implementer=data.get('implementer', ''),
            confirm_person=data.get('confirm_person'),
            acceptor=data.get('acceptor'),
            effect_evaluation=data.get('effect_evaluation'),
            evaluator=data.get('evaluator'),
            remarks=data.get('remarks'),
            shift_type=data.get('shift_type', '其他'),
            created_by=request.user
        )
        
        # 创建或获取设备
        equipment, created = Equipment.objects.get_or_create(
            code=record.equipment_code,
            defaults={
                'name': record.equipment_name,
                'production_line': record.production_line,
                'process': record.process,
                'part_details': record.equipment_part
            }
        )
        
        # 创建维修历史记录
        MaintenanceHistory.objects.create(
            equipment=equipment,
            maintenance_record=record
        )
        
        return JsonResponse({
            'message': '维修记录创建成功',
            'record_id': record.id
        })
        
    except Exception as e:
        return JsonResponse({'error': f'创建维修记录时出错: {str(e)}'}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_equipment_list(request):
    """
    获取设备列表
    """
    try:
        equipment_list = Equipment.objects.all().order_by('name')
        
        equipment_data = []
        for eq in equipment_list:
            equipment_data.append({
                'id': eq.id,
                'name': eq.name,
                'code': eq.code,
                'production_line': eq.production_line,
                'process': eq.process,
                'part_details': eq.part_details,
                'created_at': eq.created_at
            })
        
        return JsonResponse({
            'equipment': equipment_data,
            'count': len(equipment_data)
        })
        
    except Exception as e:
        return JsonResponse({'error': f'获取设备列表时出错: {str(e)}'}, status=500)