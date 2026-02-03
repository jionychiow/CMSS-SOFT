from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from rest_framework.decorators import api_view
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from .models_maintenance_new import ShiftMaintenanceRecord
from .serializers_new import ShiftMaintenanceRecordSerializer
import json


@api_view(['GET'])
def download_excel_template(request):
    """
    下载维修记录Excel模板
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "维修记录模板"
    
    # 设置标题行
    headers = [
        '设备名称', '设备编号', '生产线', '工序', '变更原因', 
        '变更前', '变更后', '开始时间', '结束时间', '持续时间(分钟)',
        '操作人', '确认人(长白班)', '接令人(倒班)', '备注', '班次类型', '期数'
    ]
    
    # 设置标题样式
    title_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题行
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = title_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # 调整列宽
    for col_num, header in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    # 添加示例数据
    sample_data = [
        ['空压机', 'KYLJ-001', '一期生产线A', '装配工序', '设备保养',
         '待保养', '已保养', '2024-01-01 08:00', '2024-01-01 09:00', '60',
         '张三', '李四', '', 'long_day_shift', 'phase_1']
    ]
    
    for row_num, row_data in enumerate(sample_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
    
    # 保存到内存流
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=维修记录模板.xlsx'
    
    return response


@api_view(['POST'])
def download_filtered_records(request):
    """
    下载过滤后的维修记录
    """
    try:
        # 从前端获取过滤条件
        filters = request.data
        phase_code = filters.get('phase', '')
        shift_type = filters.get('shift_type', '')
        month_filter = filters.get('month', '')  # 月份筛选参数，格式为 'YYYY-MM' 或 'all'
        
        # 构建查询条件
        query_params = {}
        if phase_code:
            # 查找对应的PlantPhase对象
            from .models import PlantPhase
            try:
                plant_phase = PlantPhase.objects.get(code=phase_code)
                query_params['phase'] = plant_phase
            except PlantPhase.DoesNotExist:
                # 如果找不到对应的期数，返回空结果
                response_content = json.dumps({'error': f'找不到对应的期数: {phase_code}'}, ensure_ascii=False)
                return HttpResponse(response_content, content_type='application/json', status=404)
        if shift_type:
            # 查找对应的ShiftType对象
            from .models import ShiftType
            try:
                shift_obj = ShiftType.objects.get(code=shift_type)
                query_params['shift_type'] = shift_obj
            except ShiftType.DoesNotExist:
                # 如果找不到对应的班次类型，返回空结果
                response_content = json.dumps({'error': f'找不到对应的班次类型: {shift_type}'}, ensure_ascii=False)
                return HttpResponse(response_content, content_type='application/json', status=404)
        
        # 查询记录
        records = ShiftMaintenanceRecord.objects.filter(**query_params)
        
        # 根据月份进行筛选
        if month_filter and month_filter != 'all':
            try:
                year, month = month_filter.split('-')
                year = int(year)
                month = int(month)
                from django.db.models import Q
                
                # 筛选开始时间或结束时间在指定月份的记录（同时处理非空值）
                # 因为现在系统使用北京时间，直接按年月筛选即可
                records = records.filter(
                    (Q(start_datetime__isnull=False) & 
                     Q(start_datetime__year=year, start_datetime__month=month)) |
                    (Q(end_datetime__isnull=False) & 
                     Q(end_datetime__year=year, end_datetime__month=month))
                )
                
            except ValueError:
                # 如果月份格式不正确，返回错误
                response_content = json.dumps({'error': f'月份格式不正确，请使用 YYYY-MM 格式'}, ensure_ascii=False)
                return HttpResponse(response_content, content_type='application/json', status=400)
        
        records = records.order_by('-created_at')
        
        records = records.order_by('-created_at')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "维修记录"
        
        # 定义字段映射 - 使用与模型一致的字段名，按要求的顺序排列
        field_headers = [
            ('phase', '期数'),
            ('shift_type', '班次类型'),
            ('month', '月份'),
            ('serial_number', '序号'),
            ('equipment_name', '设备名称'),
            ('equipment_number', '设备编号'),
            ('production_line', '生产线'),
            ('process', '工序'),
            ('change_reason', '变更原因'),
            ('before_change', '变更前'),
            ('after_change', '变更后'),
            ('start_datetime', '开始日期及时间'),
            ('end_datetime', '结束日期及时间'),
            ('duration_display', '耗用时长'),  # 使用属性而非字段
            ('implementer', '实施人'),
            ('confirm_person', '确认人'),
            ('acceptor', '验收人'),
            ('remarks', '备注')
        ]
        
        # 设置标题样式
        title_font = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入标题行
        headers = [header for _, header in field_headers]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = title_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # 调整列宽
        for col_num, header in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].width = 15
        
        # 写入数据行
        for row_num, record in enumerate(records, 2):
            values = []
            for field_key, _ in field_headers:
                if field_key == 'phase':
                    value = '一期' if record.phase.code == 'phase_1' else '二期'
                elif field_key == 'shift_type':
                    value = '长白班' if record.shift_type.code == 'long_day_shift' else '倒班'
                elif field_key in ['start_datetime', 'end_datetime']:
                    dt_value = getattr(record, field_key, None)
                    if dt_value:
                        # 直接使用本地时间，不再需要时区转换
                        value = dt_value.strftime('%Y-%m-%d %H:%M')
                    else:
                        value = ''
                elif field_key == 'duration_display':
                    # 计算持续时间
                    if record.start_datetime and record.end_datetime:
                        duration_delta = record.end_datetime - record.start_datetime
                        days = duration_delta.days
                        hours, remainder = divmod(duration_delta.seconds, 3600)
                        minutes = remainder // 60
                        if days > 0:
                            value = f"{days}天{hours}小时{minutes}分钟"
                        elif hours > 0:
                            value = f"{hours}小时{minutes}分钟"
                        else:
                            value = f"{minutes}分钟"
                    else:
                        value = ''
                elif field_key == 'change_reason':
                    # 处理变更原因字段，显示中文
                    reason_map = {
                        'maintenance': '维保',
                        'repair': '维修',
                        'technical_modification': '技改'
                    }
                    value = reason_map.get(getattr(record, field_key, ''), getattr(record, field_key, ''))
                elif hasattr(record, field_key):
                    value = getattr(record, field_key, '')
                    if value is None:
                        value = ''
                else:
                    value = ''
                
                values.append(value)
            
            for col_num, value in enumerate(values, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = thin_border
        
        # 保存到内存流
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=维修记录.xlsx'
        
        return response
    
    except Exception as e:
        # 输出错误信息到控制台
        import traceback
        print(f"ERROR in download_filtered_records: {str(e)}")
        traceback.print_exc()  # 打印完整的堆栈跟踪
        
        # 使用HttpResponse而不是Response，保持一致性
        import json
        response_content = json.dumps({'error': str(e)}, ensure_ascii=False)
        return HttpResponse(response_content, content_type='application/json', status=500)


@api_view(['POST'])
def upload_excel_records(request):
    """
    上传Excel文件并导入维修记录
    """
    try:
        if 'file' not in request.FILES:
            return Response({'error': '没有上传文件'}, status=400)
        
        uploaded_file = request.FILES['file']
        
        # 检查文件类型
        if not uploaded_file.name.endswith(('.xlsx', '.xls')):
            return Response({'error': '只支持Excel文件(.xlsx, .xls)'}, status=400)
        
        # 从请求数据中获取期数和班次类型（如果提供的话）
        phase_code = request.data.get('phase', '')  # 从页面上下文获取期数
        shift_type_code = request.data.get('shift_type', '')  # 从页面上下文获取班次类型
        
        import pandas as pd
        from io import BytesIO
        
        # 使用pandas读取Excel文件
        df = pd.read_excel(BytesIO(uploaded_file.read()))
        
        print(f"DEBUG: 读取到 {len(df)} 行数据")  # 调试信息
        print(f"DEBUG: 从请求中获取的期数: {phase_code}, 班次类型: {shift_type_code}")  # 调试信息
        
        imported_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                print(f"DEBUG: 处理第 {index + 2} 行数据")  # 调试信息
                
                # 尝试解析日期时间
                start_datetime = row.get('开始日期及时间')
                end_datetime = row.get('结束日期及时间')
                
                print(f"DEBUG: 原始期数值: {row.get('期数', '')}, 原始班次值: {row.get('班次类型', '')}")  # 调试信息
                
                # 优先使用从Excel读取的期数和班次类型，如果不存在则使用从请求参数获取的默认值
                excel_phase_text = row.get('期数', '')
                excel_shift_text = row.get('班次类型', '')
                
                # 如果Excel中有期数和班次信息，优先使用Excel中的；否则使用请求参数中的
                if excel_phase_text and isinstance(excel_phase_text, str) and excel_phase_text.strip():
                    # 从Excel解析期数
                    if '一' in excel_phase_text or '1' in str(excel_phase_text):
                        final_phase_code = 'phase_1'
                    elif '二' in excel_phase_text or '2' in str(excel_phase_text):
                        final_phase_code = 'phase_2'
                    else:
                        final_phase_code = ''
                else:
                    # 使用从请求参数获取的期数
                    final_phase_code = phase_code
                
                if excel_shift_text and isinstance(excel_shift_text, str) and excel_shift_text.strip():
                    # 从Excel解析班次类型
                    if '长白班' in excel_shift_text or '白班' in excel_shift_text:
                        final_shift_type_code = 'long_day_shift'
                    elif '倒班' in excel_shift_text:
                        final_shift_type_code = 'rotating_shift'
                    else:
                        final_shift_type_code = ''
                else:
                    # 使用从请求参数获取的班次类型
                    final_shift_type_code = shift_type_code
                
                print(f"DEBUG: 最终期数代码: {final_phase_code}, 最终班次类型代码: {final_shift_type_code}")  # 调试信息
                
                # 验证必要字段
                if not final_shift_type_code or not final_phase_code:
                    errors.append(f"第{index + 2}行缺少期数或班次类型信息: phase='{final_phase_code}', shift_type='{final_shift_type_code}'")
                    print(f"DEBUG: 第{index + 2}行因缺少必要字段被跳过")  # 调试信息
                    continue
                
                # 获取外键对象
                from .models import PlantPhase, ShiftType
                try:
                    phase_obj = PlantPhase.objects.get(code=final_phase_code)
                    shift_type_obj = ShiftType.objects.get(code=final_shift_type_code)
                    print(f"DEBUG: 成功找到外键对象 - phase_id: {phase_obj.id}, shift_type_id: {shift_type_obj.id}")  # 调试信息
                except PlantPhase.DoesNotExist:
                    errors.append(f"第{index + 2}行期数代码不存在: {final_phase_code}")
                    print(f"DEBUG: 期数代码不存在: {final_phase_code}")  # 调试信息
                    continue
                except ShiftType.DoesNotExist:
                    errors.append(f"第{index + 2}行班次类型代码不存在: {final_shift_type_code}")
                    print(f"DEBUG: 班次类型代码不存在: {final_shift_type_code}")  # 调试信息
                    continue
                
                # 处理日期时间格式
                if pd.notna(start_datetime):
                    start_datetime = pd.to_datetime(start_datetime)
                if pd.notna(end_datetime):
                    end_datetime = pd.to_datetime(end_datetime)
                
                # 创建或更新记录
                record_data = {
                    'equipment_name': row.get('设备名称', ''),
                    'equipment_number': row.get('设备编号', ''),
                    'production_line': row.get('生产线', ''),
                    'process': row.get('工序', ''),
                    'change_reason': row.get('变更原因', ''),
                    'before_change': row.get('变更前', ''),
                    'after_change': row.get('变更后', ''),
                    'start_datetime': start_datetime,
                    'end_datetime': end_datetime,
                    'duration': row.get('耗用时长', 0),
                    'implementer': row.get('实施人', ''),
                    'confirm_person': row.get('确认人', ''),
                    'acceptor': row.get('验收人', ''),
                    'remarks': row.get('备注', ''),
                }
                
                print(f"DEBUG: 准备保存记录，设备名称: {record_data['equipment_name']}")  # 调试信息
                
                # 保存记录 - 使用create方法，让模型自动生成序号
                record = ShiftMaintenanceRecord(
                    equipment_name=record_data['equipment_name'],
                    equipment_number=record_data['equipment_number'],
                    production_line=record_data['production_line'],
                    process=record_data['process'],
                    change_reason=record_data['change_reason'],
                    before_change=record_data['before_change'],
                    after_change=record_data['after_change'],
                    start_datetime=record_data['start_datetime'],
                    end_datetime=record_data['end_datetime'],
                    duration=record_data['duration'],
                    implementer=record_data['implementer'],
                    confirm_person=record_data['confirm_person'],
                    acceptor=record_data['acceptor'],
                    remarks=record_data['remarks'],
                    shift_type=shift_type_obj,  # 使用外键对象
                    phase=phase_obj  # 使用外键对象
                    # 序号和月份由模型自动生成
                )
                record.save()
                imported_count += 1
                print(f"DEBUG: 成功保存第 {index + 2} 行记录，总数: {imported_count}")  # 调试信息
                
            except Exception as e:
                errors.append(f"第{index + 2}行处理失败: {str(e)}")
                print(f"ERROR: 第{index + 2}行处理失败: {str(e)}")  # 调试信息
        
        return Response({
            'message': f'成功导入 {imported_count} 条记录',
            'errors': errors
        })
    
    except Exception as e:
        import traceback
        print(f"ERROR in upload_excel_records: {str(e)}")
        traceback.print_exc()  # 打印完整的堆栈跟踪
        return Response({'error': str(e)}, status=500)